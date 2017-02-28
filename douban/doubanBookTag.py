#! /usr/bin/env python
# -*- coding=utf-8 -*-


import requests
import bs4
import sys
import urllib
import urllib2
import openpyxl
import time
import numpy


reload(sys)
sys.setdefaultencoding('utf8')


hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def dou_ban_book_search(bookTagLists):
    bookLists = []
    for bookTag in bookTagLists:
        bookList = book_search(bookTag)
        # sorted the data by rating num
        bookList = sorted(bookList, key=lambda x: x[1], reverse=True)
        bookLists.append(bookList)

    return bookLists


def book_search(bookTag):
    bookList = []
    # pageIdx = 49  # test page index
    pageIdx = 0

    while True:
        print('Getting book info for tag:%s on Page: %d' % (bookTag.decode(), pageIdx + 1))
        url = 'https://book.douban.com/tag/' + urllib.quote(bookTag) \
                + '?start=' + str(pageIdx * 20) + '&type=T'
        time.sleep(numpy.random.rand() * 2)

        try:
            request = urllib2.Request(url, headers=hds[pageIdx % len(hds)])
            response = urllib2.urlopen(request).read()
            soup = bs4.BeautifulSoup(str(response))
        except (urllib2.HTTPError, urllib2.URLError) as e:
            print(e)
            continue

        # res = requests.get(url)
        # try:
        #     res.raise_for_status()
        #     soup = bs4.BeautifulSoup(res.text)
        # except Exception as e:
        #     print('There was an error for get url...')
        #     raise e

        subjectListSoup = soup.select('div[id="subject_list"]')
        subjectItem = subjectListSoup[0].select('li[class="subject-item"]')
        if len(subjectItem) == 0:
            print('We reached the last page, got %d books...' % (len(bookList)))
            break
        else:
            for bookSubjectItem in subjectItem:
                bookList.append(get_book_info(bookSubjectItem))

        pageIdx += 1
    return bookList


def get_book_info(bookSubjectItem):
    bookTitle = bookSubjectItem.select('a[title]')[0].get('title')  # maybe we will use encode()
    bookUrl = bookSubjectItem.select('a[href]')[0].get('href')
    authorInfo, pressInfo = get_book_base_info(bookSubjectItem)

    try:
        ratingNum = bookSubjectItem.select('span[class="rating_nums"]')[0].get_text().encode('utf-8')
    except:
        ratingNum = '0.0'

    return [bookTitle, ratingNum, authorInfo, pressInfo, bookUrl]


def get_book_base_info(bookSubjectItem):
    bookPublic = bookSubjectItem.select('div[class="pub"]')[0].get_text().strip()
    infoList = bookPublic.split('/')
    try:
        authorInfo = 'Author: ' + '/'.join(infoList[0:-3])
    except:
        authorInfo = 'Author: ' + 'None'

    try:
        pressInfo = '/'.join(infoList[-3:])
    except:
        pressInfo = 'None'

    return authorInfo, pressInfo


def save_book_to_excel(bookLists, bookTagLists):
    wb = openpyxl.Workbook()
    ws = []
    for tag in range(len(bookTagLists)):
        ws.append(wb.create_sheet(title=bookTagLists[tag].decode()))

    for tag in range(len(bookTagLists)):
        ws[tag].append(['Id', 'Titil', 'Rating', 'Author', 'Press', 'Link'])
        count = 1
        for bookList in bookLists[tag]:
            ws[tag].append([count, bookList[0], float(bookList[1]), bookList[2], bookList[3], bookList[4]])
            count += 1

    fileName = 'bookList'
    for i in range(len(bookTagLists)):
        fileName += ('-' + bookTagLists[i].decode())
    fileName += '.xlsx'
    wb.save(fileName)
    print('Save to file:%s successfully...' % (fileName))


def main():
    # bookTagLists = ['编程', '科学']
    # bookTagLists = ['编程', '文学', '小说']
    bookTagLists = ['王小波', '村上春树', '鲁迅', '文学', '社会', '数学', '经济学', '科普', '互联网', '科技']
    bookLists = dou_ban_book_search(bookTagLists)
    save_book_to_excel(bookLists, bookTagLists)


if __name__ == '__main__':
    main()
